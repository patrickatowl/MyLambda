import sys, io, subprocess, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

repo_file="./repo_list.txt"

def load_repo_filter(filepath):
    """Load repo names from a text file, one per line"""
    try:
        with open(filepath) as f:
            return [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        print(f"Warning: repo filter file '{filepath}' not found, loading all repos.", file=sys.stderr)
        return []
    
repos = load_repo_filter(repo_file)

results = []

# --- Step 1: 收集所有 Repo 的使用者權限資料 ---
for repo in repos:
    # 使用 collaborators API，並帶上 per_page=100 參數防止人數較多時被分頁切斷
    r = subprocess.run(
        ['gh', 'api', f'repos/owlting/{repo}/collaborators?per_page=100'],
        capture_output=True, text=True
    )
    
    if r.returncode != 0:
        err = r.stderr.strip()
        results.append({
            'repo': repo,
            'username': 'N/A',
            'role_name': 'N/A',
            'permission_admin': 'N/A',
            'permission_push': 'N/A',
            'permission_pull': 'N/A',
            'status': f'error: {err[:80]}'
        })
        print(f'SKIP {repo}: {err[:60]}')
        continue

    try:
        users = json.loads(r.stdout)
    except json.JSONDecodeError:
        results.append({
            'repo': repo,
            'username': 'N/A',
            'role_name': 'N/A',
            'permission_admin': 'N/A',
            'permission_push': 'N/A',
            'permission_pull': 'N/A',
            'status': 'error: JSON parse failed'
        })
        print(f'ERROR {repo}: Failed to parse JSON')
        continue

    # 如果該 Repo 沒有任何權限設定（通常不可能，至少會有擁有者）
    if not users:
        results.append({
            'repo': repo,
            'username': 'No Collaborators',
            'role_name': '-',
            'permission_admin': '-',
            'permission_push': '-',
            'permission_pull': '-',
            'status': 'success'
        })
        continue

    # 拆解每個使用者的權限
    for user in users:
        username = user.get('login', '')
        role_name = user.get('role_name', '') # 例如: admin, write, read
        permissions = user.get('permissions', {}) # 包含 admin, push, pull 等布林值
        
        results.append({
            'repo': repo,
            'username': username,
            'role_name': role_name,
            'permission_admin': str(permissions.get('admin', '')),
            'permission_push': str(permissions.get('push', '')),
            'permission_pull': str(permissions.get('pull', '')),
            'status': 'success'
        })
    print(f'OK {repo} (Found {len(users)} collaborators)')

# --- Step 2: 統一寫入 Excel ---
wb = Workbook()
ws = wb.active
ws.title = 'repo_collaborators'

# 定義新的 Excel 表頭
headers = [
    'repo', 'username', 'role_name', 
    'permission_admin', 'permission_push', 'permission_pull', 
    'status'
]

# 寫入標題列
header_fill = PatternFill('solid', start_color='203764') # 換個深藍色區隔
header_font = Font(bold=True, color='FFFFFF')
for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# 寫入資料列
for ri, row in enumerate(results, 2):
    for ci, h in enumerate(headers, 1):
        ws.cell(row=ri, column=ci, value=row.get(h, ''))

# 自動調整欄寬
for col in ws.columns:
    max_len = max(len(str(c.value or '')) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

# 儲存檔案
save_path = r'repo_collaborators.xlsx'
wb.save(save_path)
print(f'\nAll done! Saved to {save_path}')