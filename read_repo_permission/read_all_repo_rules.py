import sys, io, subprocess, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

repo_file="./repo_list.txt"
"""
repos = [
    'market-kit-android','market_admin','market_agent_search','market_api_doc',
    'market_edm_emails_update','market_fb_bot','market_google_tool','market_motherday_activity',
    'market_sap','market_statement','market_v2','market_v2_apidoc','market_vendor',
    'owlting_market','owlting_market_admin','owlting_market_cart_recs',
    'owlting_market_content_ai','owlting_market_flutter_app','owlting_market_frontend',
    'owlting_market_recs','prediction_market'
]
"""


def load_repo_filter(filepath):
    """Load repo names from a text file, one per line"""
    try:
        with open(filepath) as f:
            return [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        print(f"Warning: repo filter file '{filepath}' not found, loading all repos.", file=sys.stderr)
        return []
    
repos = load_repo_filter(repo_file)

results = []

# --- Step 1: 收集所有 Repo 的 API 資料 ---
for repo in repos:
    r = subprocess.run(
        ['gh', 'api', f'repos/owlting/{repo}/branches/master/protection'],
        capture_output=True, text=True
    )
    if r.returncode != 0:
        err = r.stderr.strip()
        if 'Branch not protected' in err or '404' in err:
            results.append({'repo': repo, 'status': 'no protection'})
        else:
            results.append({'repo': repo, 'status': f'error: {err[:80]}'})
        print(f'SKIP {repo}: {err[:60]}')
        continue

    try:
        d = json.loads(r.stdout)
    except json.JSONDecodeError:
        results.append({'repo': repo, 'status': 'error: JSON parse failed'})
        print(f'ERROR {repo}: Failed to parse JSON')
        continue

    rsc = d.get('required_status_checks', {})
    rpr = d.get('required_pull_request_reviews', {})
    restrict = d.get('restrictions', {})
    push_users = [u['login'] for u in restrict.get('users', [])] if restrict else []
    push_teams = [t['slug'] for t in restrict.get('teams', [])] if restrict else []

    results.append({
        'repo': repo,
        'status': 'protected',
        'strict_status_checks': rsc.get('strict', ''),
        'required_checks': ', '.join([c['context'] for c in rsc.get('checks', [])]),
        'required_reviews': rpr.get('required_approving_review_count', ''),
        'dismiss_stale_reviews': rpr.get('dismiss_stale_reviews', ''),
        'require_code_owner_reviews': rpr.get('require_code_owner_reviews', ''),
        'require_last_push_approval': rpr.get('require_last_push_approval', ''),
        'enforce_admins': d.get('enforce_admins', {}).get('enabled', ''),
        'allow_force_pushes': d.get('allow_force_pushes', {}).get('enabled', ''),
        'allow_deletions': d.get('allow_deletions', {}).get('enabled', ''),
        'block_creations': d.get('block_creations', {}).get('enabled', ''),
        'required_linear_history': d.get('required_linear_history', {}).get('enabled', ''),
        'required_conversation_resolution': d.get('required_conversation_resolution', {}).get('enabled', ''),
        'required_signatures': d.get('required_signatures', {}).get('enabled', ''),
        'push_restrictions_users': ', '.join(push_users),
        'push_restrictions_teams': ', '.join(push_teams),
    })
    print(f'OK {repo}')

# --- Step 2: 統一寫入 Excel (移到迴圈外) ---
wb = Workbook()
ws = wb.active
ws.title = 'market_branch_protection'

headers = [
    'repo', 'status', 'strict_status_checks', 'required_checks',
    'required_reviews', 'dismiss_stale_reviews', 'require_code_owner_reviews',
    'require_last_push_approval', 'enforce_admins', 'allow_force_pushes',
    'allow_deletions', 'block_creations', 'required_linear_history',
    'required_conversation_resolution', 'required_signatures',
    'push_restrictions_users', 'push_restrictions_teams'
]

# 寫入標題列
header_fill = PatternFill('solid', start_color='4472C4')
header_font = Font(bold=True, color='FFFFFF')
for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# 寫入資料列
for ri, row in enumerate(results, 2):
    for ci, h in enumerate(headers, 1):
        ws.cell(row=ri, column=ci, value=row.get(h, ''))

# 自動調整欄寬
for col in ws.columns:
    max_len = max(len(str(c.value or '')) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

# 儲存檔案（整支程式只儲存這一次）
save_path = r'branch_protection.xlsx'
wb.save(save_path)
print(f'\nAll done! Saved to {save_path}')