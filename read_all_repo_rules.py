import sys, io, subprocess, json, argparse, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

parser = argparse.ArgumentParser()
parser.add_argument('--repo-list', default='read_repo_permission/repo_list.txt', help='Path to repo list file (one repo per line)')
parser.add_argument('--org', default='owlting', help='GitHub organization name')
parser.add_argument('--branch', default='master', help='Branch name to check protection for')
parser.add_argument('--output', default='git_read_repos_jobs/branch_protection.xlsx', help='Output xlsx path')
args = parser.parse_args()

repo_list_path = args.repo_list
if not os.path.isabs(repo_list_path):
    repo_list_path = os.path.join(os.path.dirname(__file__), repo_list_path)

with open(repo_list_path, encoding='utf-8') as f:
    repos = [line.strip() for line in f if line.strip() and not line.startswith('#')]

print(f'Loaded {len(repos)} repos from {repo_list_path}')

results = []

# --- Step 1: 收集所有 Repo 的 API 資料 ---
for repo in repos:
    r = subprocess.run(
        ['gh', 'api', f'repos/{args.org}/{repo}/branches/{args.branch}/protection'],
        capture_output=True, text=True
    )
    if r.returncode != 0:
        err = r.stderr.strip()
        if 'Branch not protected' in err or '404' in err:
            results.append({'repo': repo, 'status': 'no protection'})
        else:
            results.append({'repo': repo, 'status': f'error: {err[:80]}'})
        print(f'SKIP {repo}: {err[:60]}')
        continue

    try:
        d = json.loads(r.stdout)
    except json.JSONDecodeError:
        results.append({'repo': repo, 'status': 'error: JSON parse failed'})
        print(f'ERROR {repo}: Failed to parse JSON')
        continue

    rsc = d.get('required_status_checks', {})
    rpr = d.get('required_pull_request_reviews', {})
    restrict = d.get('restrictions', {})
    push_users = [u['login'] for u in restrict.get('users', [])] if restrict else []
    push_teams = [t['slug'] for t in restrict.get('teams', [])] if restrict else []

    results.append({
        'repo': repo,
        'status': 'protected',
        'strict_status_checks': rsc.get('strict', ''),
        'required_checks': ', '.join([c['context'] for c in rsc.get('checks', [])]),
        'required_reviews': rpr.get('required_approving_review_count', ''),
        'dismiss_stale_reviews': rpr.get('dismiss_stale_reviews', ''),
        'require_code_owner_reviews': rpr.get('require_code_owner_reviews', ''),
        'require_last_push_approval': rpr.get('require_last_push_approval', ''),
        'enforce_admins': d.get('enforce_admins', {}).get('enabled', ''),
        'allow_force_pushes': d.get('allow_force_pushes', {}).get('enabled', ''),
        'allow_deletions': d.get('allow_deletions', {}).get('enabled', ''),
        'block_creations': d.get('block_creations', {}).get('enabled', ''),
        'required_linear_history': d.get('required_linear_history', {}).get('enabled', ''),
        'required_conversation_resolution': d.get('required_conversation_resolution', {}).get('enabled', ''),
        'required_signatures': d.get('required_signatures', {}).get('enabled', ''),
        'push_restrictions_users': ', '.join(push_users),
        'push_restrictions_teams': ', '.join(push_teams),
    })
    print(f'OK {repo}')

# --- Step 2: 統一寫入 Excel (移到迴圈外) ---
wb = Workbook()
ws = wb.active
ws.title = 'branch_protection'

headers = [
    'repo', 'status', 'strict_status_checks', 'required_checks',
    'required_reviews', 'dismiss_stale_reviews', 'require_code_owner_reviews',
    'require_last_push_approval', 'enforce_admins', 'allow_force_pushes',
    'allow_deletions', 'block_creations', 'required_linear_history',
    'required_conversation_resolution', 'required_signatures',
    'push_restrictions_users', 'push_restrictions_teams'
]

# 寫入標題列
header_fill = PatternFill('solid', start_color='4472C4')
header_font = Font(bold=True, color='FFFFFF')
for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# 寫入資料列
for ri, row in enumerate(results, 2):
    for ci, h in enumerate(headers, 1):
        ws.cell(row=ri, column=ci, value=row.get(h, ''))

# 自動調整欄寬
for col in ws.columns:
    max_len = max(len(str(c.value or '')) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

# 儲存檔案（整支程式只儲存這一次）
save_path = args.output
if not os.path.isabs(save_path):
    save_path = os.path.join(os.path.dirname(__file__), save_path)
os.makedirs(os.path.dirname(save_path), exist_ok=True)
wb.save(save_path)
print(f'\nAll done! Saved to {save_path}')